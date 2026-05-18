"""
ollama_council_batch_v2.py
--------------------------
Improvements over v1:
  - Saves a row to Excel after EVERY question (no crash = no data loss)
  - Resume mode: skips question IDs already present in the output file
  - Summary includes mean, median, stdev, N per chairman model
  - Batch-safe: run with --limit 20 repeatedly; duplicates are skipped
"""

import argparse
import json
import math
import statistics
import time
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
import requests
from openpyxl import load_workbook, Workbook

OLLAMA_URL = "http://localhost:11434/api/chat"

# ── System prompts ────────────────────────────────────────────────────────────

ANSWER_SYSTEM = (
    "You are helping with a research benchmark of timely prediction questions. "
    "Answer as if the outcome is not yet known. Be concise and structured."
)

COUNCIL_SYSTEM = (
    "You are the chairman of an LLM council. You will read answers from multiple models, "
    "compare them, and produce one best final answer. Prefer answers that are consistent, "
    "well-supported, and aligned with the allowed answer choices."
)

JUDGE_SYSTEM = (
    "You are an evaluator for an unresolved prediction question. "
    "Your job is to score answer quality, not real-world correctness, because the outcome may not be known yet. "
    "Focus on clarity, reasoning quality, relevance, and whether the final answer matches one of the allowed choices."
)

# ── Prompt templates ──────────────────────────────────────────────────────────

ANSWER_PROMPT_TEMPLATE = (
    "Question: {question}\n\n"
    "Allowed answer choices: {choices}\n\n"
    "Return exactly these two sections:\n"
    "1. Predicted answer: <one final answer from the allowed answer choices if possible>\n"
    "2. Reasoning: <2-4 sentences>\n"
)

COUNCIL_PROMPT_TEMPLATE = (
    "Question: {question}\n\n"
    "Allowed answer choices: {choices}\n\n"
    "Model answers:\n{answers_block}\n\n"
    "Produce exactly these two sections:\n"
    "1. Council answer: <one final answer from the allowed answer choices if possible>\n"
    "2. Council reasoning: <2-5 sentences explaining which answer is strongest and why>\n"
)

JUDGE_PROMPT_TEMPLATE = (
    "Question: {question}\n\n"
    "Allowed answer choices: {choices}\n\n"
    "Answer being judged:\n{candidate_answer}\n\n"
    "Reasoning being judged:\n{candidate_reasoning}\n\n"
    "Score this response from 1 to 5 on:\n"
    "- Clarity\n"
    "- Reasoning quality\n"
    "- Relevance\n"
    "- Overall quality\n\n"
    "Use whole numbers only.\n\n"
    "Return valid JSON exactly in this schema:\n"
    '{{\n'
    '  "clarity": <1-5>,\n'
    '  "reasoning_quality": <1-5>,\n'
    '  "relevance": <1-5>,\n'
    '  "overall": <1-5>,\n'
    '  "note": "<one short sentence>"\n'
    '}}\n'
)

# ── Ollama call ───────────────────────────────────────────────────────────────

def call_ollama(model: str, system: str, prompt: str,
                temperature: float = 0.2, timeout: int = 120) -> str:
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system},
            {"role": "user",   "content": prompt},
        ],
        "stream": False,
        "options": {"temperature": temperature},
    }
    r = requests.post(OLLAMA_URL, json=payload, timeout=timeout)
    r.raise_for_status()
    return r.json()["message"]["content"].strip()

# ── Parsing helpers ───────────────────────────────────────────────────────────

def extract_field(text: str, prefixes: List[str]) -> str:
    lower = text.lower()
    for prefix in prefixes:
        idx = lower.find(prefix.lower())
        if idx != -1:
            after = text[idx + len(prefix):].strip()
            lines = after.splitlines()
            return lines[0].strip(" -") if lines else after.strip()
    return text.splitlines()[0].strip()


def parse_answer_response(text: str) -> Dict[str, str]:
    pred   = extract_field(text, ["Predicted answer:", "Answer:"])
    lower  = text.lower()
    idx    = lower.find("reasoning:")
    reason = text[idx + len("reasoning:"):].strip() if idx != -1 else (
        " ".join(p.strip() for p in text.splitlines()[1:] if p.strip())
    )
    return {"predicted_answer": pred, "reasoning": reason or text}


def parse_council_response(text: str) -> Dict[str, str]:
    ans   = extract_field(text, ["Council answer:", "Predicted answer:", "Answer:"])
    lower = text.lower()
    idx   = lower.find("council reasoning:")
    if idx != -1:
        reason = text[idx + len("council reasoning:"):].strip()
    else:
        idx = lower.find("reasoning:")
        reason = text[idx + len("reasoning:"):].strip() if idx != -1 else text
    return {"council_answer": ans, "council_reasoning": reason}


def parse_judge_json(text: str) -> Dict[str, Any]:
    try:
        return json.loads(text)
    except Exception:
        pass
    s, e = text.find("{"), text.rfind("}")
    if s != -1 and e > s:
        try:
            return json.loads(text[s:e+1])
        except Exception:
            pass
    out = {"clarity": None, "reasoning_quality": None,
           "relevance": None, "overall": None, "note": text[:200]}
    for key in ["clarity", "reasoning_quality", "relevance", "overall"]:
        pos = text.lower().find(key.replace("_", " "))
        if pos != -1:
            snippet = text[pos:pos+40]
            digits  = "".join(ch for ch in snippet if ch.isdigit())
            if digits:
                out[key] = int(digits[0])
    return out

# ── Stats helpers ─────────────────────────────────────────────────────────────

def safe_stat(fn, vals: List[float]) -> Optional[float]:
    vals = [v for v in vals if v is not None and not math.isnan(v)]
    try:
        return fn(vals) if vals else None
    except Exception:
        return None

# ── Incremental Excel writer ──────────────────────────────────────────────────

DETAILED_COLS = [
    "question_id", "question_text", "choices",
    "chairman_model", "judge_model",
    "council_answer", "council_reasoning",
    "judge_clarity", "judge_reasoning_quality", "judge_relevance", "judge_overall",
    "judge_note", "raw_council", "raw_judge",
    "answer_models", "individual_answers_json",
]

def init_detailed_xlsx(path: str) -> None:
    """Create a fresh detailed file with headers if it does not exist."""
    p = Path(path)
    if not p.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(DETAILED_COLS)
        wb.save(path)


def load_done_ids(path: str) -> set:
    """Return set of (question_id, chairman_model) already written."""
    p = Path(path)
    if not p.exists():
        return set()
    try:
        df = pd.read_excel(path)
        if "question_id" in df.columns and "chairman_model" in df.columns:
            return set(
                zip(df["question_id"].astype(str), df["chairman_model"].astype(str))
            )
    except Exception:
        pass
    return set()


def append_row_to_xlsx(path: str, row: Dict[str, Any]) -> None:
    """Append one row dict to an existing Excel file."""
    wb = load_workbook(path)
    ws = wb.active
    values = [row.get(c) for c in DETAILED_COLS]
    ws.append(values)
    wb.save(path)

# ── Per-question processing ───────────────────────────────────────────────────

def normalize_choices(v: Any) -> str:
    if pd.isna(v):
        return ""
    return str(v).strip()


def process_question(
    row: pd.Series,
    answer_models: List[str],
    chairman_models: List[str],
    judge_model: str,
    pause_s: float,
    done_pairs: set,
    detailed_path: str,
) -> int:
    """Process one question; writes rows incrementally. Returns number of rows added."""
    qid      = str(row.get("id", row.get("question_id", "?")))
    question = str(row.get("question_text", "")).strip()
    choices  = normalize_choices(row.get("answer_type", ""))
    if not question:
        return 0

    # Collect individual model answers
    answer_records: List[Dict[str, Any]] = []
    for model in answer_models:
        prompt = ANSWER_PROMPT_TEMPLATE.format(question=question, choices=choices)
        raw    = call_ollama(model, ANSWER_SYSTEM, prompt)
        parsed = parse_answer_response(raw)
        answer_records.append({
            "question_id":      qid,
            "question_text":    question,
            "choices":          choices,
            "answer_model":     model,
            "predicted_answer": parsed["predicted_answer"],
            "model_reasoning":  parsed["reasoning"],
            "raw_answer":       raw,
        })
        time.sleep(pause_s)

    answers_block = "\n\n".join(
        f"Model: {r['answer_model']}\n"
        f"Predicted answer: {r['predicted_answer']}\n"
        f"Reasoning: {r['model_reasoning']}"
        for r in answer_records
    )

    added = 0
    for chairman in chairman_models:
        pair = (qid, chairman)
        if pair in done_pairs:
            print(f"    skip (already done): q={qid}, chairman={chairman}")
            continue

        # Council
        council_prompt = COUNCIL_PROMPT_TEMPLATE.format(
            question=question, choices=choices, answers_block=answers_block
        )
        council_raw    = call_ollama(chairman, COUNCIL_SYSTEM, council_prompt)
        council_parsed = parse_council_response(council_raw)
        time.sleep(pause_s)

        # Judge
        judge_prompt = JUDGE_PROMPT_TEMPLATE.format(
            question=question,
            choices=choices,
            candidate_answer=council_parsed["council_answer"],
            candidate_reasoning=council_parsed["council_reasoning"],
        )
        judge_raw = call_ollama(judge_model, JUDGE_SYSTEM, judge_prompt)
        judge     = parse_judge_json(judge_raw)
        time.sleep(pause_s)

        out = {
            "question_id":            qid,
            "question_text":          question,
            "choices":                choices,
            "chairman_model":         chairman,
            "judge_model":            judge_model,
            "council_answer":         council_parsed["council_answer"],
            "council_reasoning":      council_parsed["council_reasoning"],
            "judge_clarity":          judge.get("clarity"),
            "judge_reasoning_quality":judge.get("reasoning_quality"),
            "judge_relevance":        judge.get("relevance"),
            "judge_overall":          judge.get("overall"),
            "judge_note":             judge.get("note"),
            "raw_council":            council_raw,
            "raw_judge":              judge_raw,
            "answer_models":          ", ".join(answer_models),
            "individual_answers_json":json.dumps(answer_records, ensure_ascii=False),
        }

        # ── INCREMENTAL SAVE ──────────────────────────────────────────────────
        append_row_to_xlsx(detailed_path, out)
        done_pairs.add(pair)
        added += 1
        print(f"    saved: q={qid}, chairman={chairman}, overall={judge.get('overall')}")

    return added

# ── Summary writer ────────────────────────────────────────────────────────────

def write_summary(detailed_path: str, summary_path: str) -> None:
    try:
        df = pd.read_excel(detailed_path)
    except Exception as e:
        print(f"Could not read {detailed_path} for summary: {e}")
        return

    if df.empty or "chairman_model" not in df.columns:
        print("No data to summarise yet.")
        return

    rows = []
    for chairman, g in df.dropna(subset=["chairman_model"]).groupby("chairman_model"):
        vals = pd.to_numeric(g["judge_overall"], errors="coerce").dropna().tolist()
        rows.append({
            "chairman_model":      chairman,
            "n":                   len(vals),
            "mean_judge_overall":  safe_stat(statistics.mean, vals),
            "median_judge_overall":safe_stat(statistics.median, vals),
            "stdev_judge_overall": safe_stat(statistics.stdev, vals) if len(vals) >= 2 else 0.0,
        })

    summary_df = pd.DataFrame(rows)
    summary_df.to_excel(summary_path, index=False)
    print(f"\n── Summary ({summary_path}) ──")
    print(summary_df.to_string(index=False))

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(
        description="Run local Ollama LLM-as-council + judge on a question dataset."
    )
    ap.add_argument("--input",           required=True,
                    help="Excel or CSV file with question_text and answer_type columns")
    ap.add_argument("--out-prefix",      default="council_results",
                    help="Prefix for output files")
    ap.add_argument("--limit",           type=int,   default=20,
                    help="Max new questions to process this run")
    ap.add_argument("--start-row",       type=int,   default=0,
                    help="0-based start offset in the dataset")
    ap.add_argument("--answer-models",   nargs="+",  default=["llama3.1:8b", "mistral"],
                    help="Ollama models used to answer")
    ap.add_argument("--chairman-models", nargs="+",  default=["llama3.1:8b", "mistral"],
                    help="Chairman model(s) to synthesise council answer")
    ap.add_argument("--judge-model",     default="llama3.1:8b",
                    help="Ollama model used as judge")
    ap.add_argument("--pause",           type=float, default=0.25,
                    help="Pause between calls (seconds)")
    args = ap.parse_args()

    # Output paths
    detailed_path = f"{args.out_prefix}_detailed.xlsx"
    summary_path  = f"{args.out_prefix}_summary.xlsx"

    # Load input
    path = Path(args.input)
    df   = pd.read_excel(path) if path.suffix.lower() in [".xlsx", ".xls"] else pd.read_csv(path)

    if "question_text" not in df.columns:
        raise ValueError("Input file must have a 'question_text' column")
    if "answer_type" not in df.columns:
        df["answer_type"] = ""
    if "id" not in df.columns and "question_id" not in df.columns:
        df["question_id"] = range(1, len(df) + 1)

    df = df[df["question_text"].notna()].copy()
    df = df[df["question_text"].astype(str).str.strip() != ""]
    df = df.iloc[args.start_row: args.start_row + args.limit]

    print(f"Input rows selected: {len(df)}")
    print(f"Answer models  : {args.answer_models}")
    print(f"Chairman models: {args.chairman_models}")
    print(f"Judge model    : {args.judge_model}")
    print(f"Output         : {detailed_path}\n")

    # Initialise / resume
    init_detailed_xlsx(detailed_path)
    done_pairs = load_done_ids(detailed_path)
    print(f"Already done (q_id, chairman) pairs loaded: {len(done_pairs)}\n")

    total_added = 0
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        qid = row.get("id", row.get("question_id", "?"))
        print(f"[{i}/{len(df)}] question_id={qid}")
        try:
            n = process_question(
                row, args.answer_models, args.chairman_models,
                args.judge_model, args.pause, done_pairs, detailed_path,
            )
            total_added += n
        except Exception as e:
            print(f"  ERROR on q={qid}: {e}")
            # Write error row so we have a record
            error_row = {c: None for c in DETAILED_COLS}
            error_row.update({
                "question_id":   str(qid),
                "question_text": str(row.get("question_text", "")),
                "choices":       str(row.get("answer_type", "")),
                "judge_note":    f"ERROR: {e}",
                "answer_models": ", ".join(args.answer_models),
            })
            append_row_to_xlsx(detailed_path, error_row)

    print(f"\n✅ Finished.  New rows written this run: {total_added}")
    write_summary(detailed_path, summary_path)


if __name__ == "__main__":
    main()
