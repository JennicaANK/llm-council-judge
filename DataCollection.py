import argparse
import json
import time
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

import requests
from dateutil import parser as dtparser
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


GAMMA_EVENTS_URL = "https://gamma-api.polymarket.com/events"


def iso(dt: datetime) -> str:
    """RFC3339-ish string Polymarket expects."""
    return dt.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")


def safe_parse_dt(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        return dtparser.isoparse(s)
    except Exception:
        return None


def parse_outcomes(outcomes_field: Any) -> List[str]:
    """
    The API returns outcomes as a string in many cases (often JSON-encoded).
    We try to convert it into a list of outcome labels.
    """
    if outcomes_field is None:
        return []
    if isinstance(outcomes_field, list):
        return [str(x) for x in outcomes_field]

    if isinstance(outcomes_field, str):
        s = outcomes_field.strip()
        # Try JSON list
        try:
            val = json.loads(s)
            if isinstance(val, list):
                return [str(x) for x in val]
        except Exception:
            pass

        # Fallback: attempt to split common delimiter patterns
        # (This is imperfect but better than empty)
        for delim in ["|", ",", ";"]:
            if delim in s and len(s) < 400:
                parts = [p.strip().strip('"').strip("'") for p in s.split(delim)]
                parts = [p for p in parts if p]
                if len(parts) >= 2:
                    return parts

    return []


def outcomes_to_answer_type(outcomes: List[str]) -> str:
    if not outcomes:
        return ""
    # Match your format: "(A, B, C)"
    return "(" + ", ".join(outcomes) + ")"


def build_event_link(event_slug: Optional[str]) -> str:
    if not event_slug:
        return ""
    return f"https://polymarket.com/event/{event_slug}"


def fetch_events_within_window(
    start_utc: datetime,
    end_utc: datetime,
    limit: int = 100,
    sleep_s: float = 0.2,
) -> List[Dict[str, Any]]:
    """
    Pull events that are active, not closed, and end within [start_utc, end_utc].
    Uses limit/offset pagination.
    """
    events: List[Dict[str, Any]] = []
    offset = 0

    while True:
        params = {
            "active": "true",
            "closed": "false",
            "limit": str(limit),
            "offset": str(offset),
            # Date window filters supported by the Events list endpoint
            "end_date_min": iso(start_utc),
            "end_date_max": iso(end_utc),
            # Order by end_date so we prioritize soonest-resolving
            "order": "end_date",
            "ascending": "true",
        }

        r = requests.get(GAMMA_EVENTS_URL, params=params, timeout=30)
        if r.status_code != 200:
            raise RuntimeError(f"Gamma API error {r.status_code}: {r.text[:300]}")

        batch = r.json()
        if not isinstance(batch, list) or len(batch) == 0:
            break

        events.extend(batch)
        offset += limit
        time.sleep(sleep_s)

    return events


def get_sheet(ws: Worksheet) -> Tuple[int, Dict[str, int]]:
    """
    Return (next_row, col_index_map) assuming the first row is headers.
    Finds the first real empty row based on question_link/question_text,
    instead of using ws.max_row which counts formatted blank rows.
    """
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h is not None}

    # Find first row where both question_link and question_text are empty
    next_row = 2
    while True:
        link_val = ws.cell(next_row, col_map["question_link"]).value
        q_val = ws.cell(next_row, col_map["question_text"]).value

        link_empty = link_val is None or str(link_val).strip() == ""
        q_empty = q_val is None or str(q_val).strip() == ""

        if link_empty and q_empty:
            break

        next_row += 1

    return next_row, col_map


def find_max_id(ws: Worksheet, id_col: int, q_col: int, link_col: int) -> int:
    """
    Find max ID only from real rows, not junk rows that only have an id.
    """
    max_id = 0
    for r in range(2, ws.max_row + 1):
        q_val = ws.cell(r, q_col).value
        link_val = ws.cell(r, link_col).value

        q_empty = q_val is None or str(q_val).strip() == ""
        link_empty = link_val is None or str(link_val).strip() == ""

        # skip fake blank rows
        if q_empty and link_empty:
            continue

        v = ws.cell(r, id_col).value
        try:
            if v is not None and str(v).strip() != "":
                max_id = max(max_id, int(v))
        except Exception:
            continue

    return max_id


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True, help="Path to your Excel template .xlsx")
    ap.add_argument("--out", required=True, help="Output .xlsx to write")
    ap.add_argument("--target", type=int, default=100, help="How many new rows to add")
    ap.add_argument("--days", type=int, default=7, help="Collect events resolving within next N days")
    ap.add_argument("--category", default="", help="Optional: only keep events matching this category (exact string)")
    args = ap.parse_args()

    # Time window: "within this week" -> next N days (default 7)
    now = datetime.now(timezone.utc)
    end = now + timedelta(days=args.days)

    # Fetch events in that window
    events = fetch_events_within_window(now, end, limit=100)

    # Load template workbook
    wb = load_workbook(args.template)
    ws = wb.active  # Sheet1
    next_row, col = get_sheet(ws)

    required_cols = ["id", "question_link", "question_text", "category", "date_collected", "resolution_date", "correct_answer", "answer_type"]
    missing = [c for c in required_cols if c not in col]
    if missing:
        raise RuntimeError(f"Template is missing columns: {missing}. Found: {list(col.keys())}")

    # Determine starting id
    start_id = find_max_id(ws, col["id"], col["question_text"], col["question_link"]) + 1

    # Avoid duplicates by link or question text already present
    existing_links = set()
    existing_q = set()
    for r in range(2, ws.max_row + 1):
        link = ws.cell(r, col["question_link"]).value
        qtxt = ws.cell(r, col["question_text"]).value
        if link:
            existing_links.add(str(link).strip())
        if qtxt:
            existing_q.add(str(qtxt).strip())

    added = 0
    curr_id = start_id
    collected_date = datetime.now().date()

    # Flatten events -> markets -> rows
    for ev in events:
        ev_category = (ev.get("category") or "").strip() or "Uncategorized"
        if args.category and ev_category != args.category:
            continue

        ev_slug = ev.get("slug")
        ev_end = safe_parse_dt(ev.get("endDate"))
        if not ev_end:
            continue

        markets = ev.get("markets") or []
        # Some events have multiple markets; we take each market's question as one row.
        for m in markets:
            if added >= args.target:
                break

            q = (m.get("question") or ev.get("title") or "").strip()
            if not q:
                continue

            link = build_event_link(ev_slug)
            if link in existing_links or q in existing_q:
                continue

            outcomes = parse_outcomes(m.get("outcomes"))
            answer_type = outcomes_to_answer_type(outcomes)

            ws.cell(next_row, col["id"]).value = curr_id
            ws.cell(next_row, col["question_link"]).value = link
            ws.cell(next_row, col["question_text"]).value = q
            ws.cell(next_row, col["category"]).value = ev_category
            ws.cell(next_row, col["date_collected"]).value = collected_date.isoformat()
            ws.cell(next_row, col["resolution_date"]).value = ev_end.date().isoformat()
            ws.cell(next_row, col["correct_answer"]).value = ""  # unresolved
            ws.cell(next_row, col["answer_type"]).value = answer_type

            existing_links.add(link)
            existing_q.add(q)

            next_row += 1
            curr_id += 1
            added += 1

        if added >= args.target:
            break

    wb.save(args.out)
    print(f"✅ Added {added} new rows to {args.out}")
    if added < args.target:
        print("Note: fewer rows than target were found within the time window. Try increasing --days or removing --category.")


if __name__ == "__main__":
    main()